# Reading an excel file using Python
from openpyxl import Workbook, load_workbook

x_coord = []
y_coord = []
z_rotation = []
w_rotation = []

_n110 = []
_n90 = []
_n60 = []
_n30 = [] 
_p0 = []
_p30= []
_p60 = []
_p90 = []
_p110 = []
 
# Give the location of the file
loc = ("/home/paulo/Desktop/take_from_bag/pose_info.xlsx")
loc2 = ("/home/paulo/Desktop/take_from_bag/scan_info.xlsx")

# To open Workbook
wb = load_workbook(loc)
sheet = wb.active

wb2 = load_workbook(loc2)
sheet2 = wb2.active

for i in range(2, sheet2.max_row):
    if ((sheet.cell(i,1).value - sheet2.cell(i,1).value) < 0.01):
        x_coord.append(sheet.cell(i,2).value)
        y_coord.append(sheet.cell(i,3).value)
        z_rotation.append(sheet.cell(i,4).value)
        w_rotation.append(sheet.cell(i,5).value)
        _n110.append(sheet2.cell(i,2).value)
        _n90.append(sheet2.cell(i,3).value)
        _n60.append(sheet2.cell(i,4).value)
        _n30.append(sheet2.cell(i,5).value)
        _p0.append(sheet2.cell(i,6).value)
        _p30.append(sheet2.cell(i,7).value)
        _p60.append(sheet2.cell(i,8).value)
        _p90.append(sheet2.cell(i,9).value)
        _p110.append(sheet2.cell(i,10).value)
    

print(len(x_coord))
print(len(y_coord))
print(len(z_rotation))
print(len(w_rotation))
print(len(_n110))
print(len(_n90))
print(len(_n60))
print(len(_n30))
print(len(_p0))
print(len(_p30))
print(len(_p60))
print(len(_p90))
print(len(_p110))
print("done")