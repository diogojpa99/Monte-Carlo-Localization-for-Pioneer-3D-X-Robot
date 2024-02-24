# Reading an excel file using Python
from openpyxl import Workbook, load_workbook
import numpy as np
import matplotlib.pyplot as plt

n_rows = 28

data1 = np.empty([n_rows,4])
data2 = np.empty([n_rows,4])
data3 = np.empty([n_rows,4])
data4 = np.empty([n_rows,4])
data5 = np.empty([n_rows,4])
data6 = np.empty([n_rows,4])
data7 = np.empty([n_rows,4])
data8 = np.empty([n_rows,4])
data9 = np.empty([n_rows,4])
data10 = np.empty([n_rows,4])


# Give the location of the file
loc1 = ("/home/diogo/Desktop/IST/SAut/Bag4_M900/1.xlsx")
loc2 = ("/home/diogo/Desktop/IST/SAut/Bag4_M900/2.xlsx")
loc3 = ("/home/diogo/Desktop/IST/SAut/Bag4_M900/3.xlsx")
loc4 = ("/home/diogo/Desktop/IST/SAut/Bag4_M900/4.xlsx")
loc5 = ("/home/diogo/Desktop/IST/SAut/Bag4_M900/5.xlsx")
loc6 = ("/home/diogo/Desktop/IST/SAut/Bag4_M900/6.xlsx")
loc7 = ("/home/diogo/Desktop/IST/SAut/Bag4_M900/7.xlsx")
loc8 = ("/home/diogo/Desktop/IST/SAut/Bag4_M900/8.xlsx")
loc9 = ("/home/diogo/Desktop/IST/SAut/Bag4_M900/9.xlsx")
loc10 = ("/home/diogo/Desktop/IST/SAut/Bag4_M900/10.xlsx")


# To open Workbook
wb1 = load_workbook(loc1)
sheet1 = wb1.active

wb2 = load_workbook(loc2)
sheet2 = wb2.active

wb3 = load_workbook(loc3)
sheet3 = wb3.active

wb4 = load_workbook(loc4)
sheet4 = wb4.active

wb5 = load_workbook(loc5)
sheet5 = wb5.active

wb6 = load_workbook(loc6)
sheet6 = wb6.active

wb7 = load_workbook(loc7)
sheet7 = wb7.active

wb8 = load_workbook(loc8)
sheet8 = wb8.active

wb9 = load_workbook(loc9)
sheet9 = wb9.active

wb10 = load_workbook(loc10)
sheet10 = wb10.active

for i in range(2, sheet1.max_row):

    data1[i][0] = sheet1.cell(i,1).value
    data1[i][1] = sheet1.cell(i,2).value
    data1[i][2] = sheet1.cell(i,3).value
    data1[i][3] = sheet1.cell(i,4).value

    data2[i][0] = sheet2.cell(i,1).value
    data2[i][1] = sheet2.cell(i,2).value
    data2[i][2] = sheet2.cell(i,3).value
    data2[i][3] = sheet2.cell(i,4).value

    data3[i][0] = sheet3.cell(i,1).value
    data3[i][1] = sheet3.cell(i,2).value
    data3[i][2] = sheet3.cell(i,3).value
    data3[i][3] = sheet3.cell(i,4).value

    data4[i][0] = sheet4.cell(i,1).value
    data4[i][1] = sheet4.cell(i,2).value
    data4[i][2] = sheet4.cell(i,3).value
    data4[i][3] = sheet4.cell(i,4).value

    data5[i][0] = sheet5.cell(i,1).value
    data5[i][1] = sheet5.cell(i,2).value
    data5[i][2] = sheet5.cell(i,3).value
    data5[i][3] = sheet5.cell(i,4).value

    data6[i][0] = sheet6.cell(i,1).value
    data6[i][1] = sheet6.cell(i,2).value
    data6[i][2] = sheet6.cell(i,3).value
    data6[i][3] = sheet6.cell(i,4).value

    data7[i][0] = sheet7.cell(i,1).value
    data7[i][1] = sheet7.cell(i,2).value
    data7[i][2] = sheet7.cell(i,3).value
    data7[i][3] = sheet7.cell(i,4).value

    data8[i][0] = sheet8.cell(i,1).value
    data8[i][1] = sheet8.cell(i,2).value
    data8[i][2] = sheet8.cell(i,3).value
    data8[i][3] = sheet8.cell(i,4).value

    data9[i][0] = sheet9.cell(i,1).value
    data9[i][1] = sheet9.cell(i,2).value
    data9[i][2] = sheet9.cell(i,3).value
    data9[i][3] = sheet9.cell(i,4).value

    data10[i][0] = sheet10.cell(i,1).value
    data10[i][1] = sheet10.cell(i,2).value
    data10[i][2] = sheet10.cell(i,3).value
    data10[i][3] = sheet10.cell(i,4).value


data_avg = np.empty([n_rows,4])
data_avg.fill(0.)

for i in range(n_rows):
        data_avg[i][0] = (data1[i][0] + data2[i][0] + data3[i][0] + data4[i][0] +data5[i][0] +data6[i][0] +data7[i][0] +data8[i][0] +data9[i][0] +data10[i][0])/10
        data_avg[i][1] = (data1[i][1] + data2[i][1] + data3[i][1] + data4[i][1] +data5[i][1] +data6[i][1] +data7[i][1] +data8[i][1] +data9[i][1] +data10[i][1])/10
        data_avg[i][2] = (data1[i][2] + data2[i][2] + data3[i][2] + data4[i][2] +data5[i][2] +data6[i][2] +data7[i][2] +data8[i][2] +data9[i][2] +data10[i][2])/10 
        data_avg[i][3] = (data1[i][3] + data2[i][3] + data3[i][3] + data4[i][3] +data5[i][3] +data6[i][3] +data7[i][3] +data8[i][3] +data9[i][3] +data10[i][3])/10 


#plot
for i in range(n_rows-1):
    plt.plot((data_avg[i][3],data_avg[i][0]) , (data_avg[i+1][3],data_avg[i+1][0]), c = 'blue')

plt.show()
